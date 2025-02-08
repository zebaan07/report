import pandas as pd
import streamlit as st
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

# Function to load recommendations from sample file
def load_recommendations(sample_file):
    df = pd.read_excel(sample_file, engine='openpyxl')
    return dict(zip(df["recommendationDisplayName"], df["owner"]))

# Function to format Excel file
def format_excel(output_file):
    wb = load_workbook(output_file)
    sheet = wb.active
    
    # Apply table formatting
    table = Table(displayName="RecommendationsTable", ref=sheet.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Apply header formatting
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4B5320", end_color="4B5320", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply alternating row colors
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        fill_color = "DCE6F1" if i % 2 == 0 else "C5D9F1"
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    wb.save(output_file)

# Function to update owner column in Excel
def update_owner_column(input_file, coe_mapping):
    df = pd.read_excel(input_file, engine='openpyxl')
    df["owner"] = df["recommendationDisplayName"].map(coe_mapping)
    df.sort_values(by=["owner"], ascending=True, inplace=True)  # Sort by owner A-Z
    output_file = "updated_" + os.path.basename(input_file)
    df.to_excel(output_file, index=False, engine='openpyxl')
    
    format_excel(output_file)
    
    return output_file

# Streamlit Web UI
st.title("Excel Owner Updater")
st.write("Developed by Zebaan")

uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])
sample_file = "recommendation_sample.xlsx"

if uploaded_file is not None:
    st.write("Processing file...")
    if os.path.exists(sample_file):
        coe_mapping = load_recommendations(sample_file)
        output_file = update_owner_column(uploaded_file, coe_mapping)
        
        with open(output_file, "rb") as file:
            btn = st.download_button(label="Download Updated Excel", data=file, file_name=output_file)
    else:
        st.error("Recommendation sample file not found!")
